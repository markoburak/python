from  selenium import webdriver
import openpyxl
import time
from datetime import datetime
from threading import Timer


def korona():
    wb = openpyxl.load_workbook('korona.xlsx')
    # open first sheet
    sheet = wb['Sheet1']


    url  = r'https://index.minfin.com.ua/ua/reference/coronavirus/geography/'
    # path to chromedriver
    chrome_path = r'E:\завантаження\chromedriver_win32\chromedriver.exe'

    driver = webdriver.Chrome(chrome_path)

    driver.get(url)

    # korona_table = driver.find_element_by_xpath("//table[@class ='line']")
    # print(korona_table.text)

    arr_table = driver.find_elements_by_xpath("//td[@class ='larger']")
    amount =[]
    '''
    country =[]
    for i in range(2,5):
        loc_country = driver.find_element_by_xpath('//*[@id="tm-table"]/div/table/tbody/tr['+str(i)+']/td[1]')
        country.append(loc_country.text)
    '''
    for elements in arr_table:
        amount.append(int(elements.text))
    try:
        amount.append(sum(int(amount)))
    except:
        pass

    all_country = driver.find_elements_by_xpath('//*[@id="tm-table"]/div/table/tbody/tr/td[1]')
    all_country_arr =[]
    for elem in all_country:
        all_country_arr.append(elem.text)
    # print(all_country_arr)
    # print(amount)

    # place team to excel
    rows = 2
    cols = 1
    for countr in all_country_arr:
        sheet.cell(row=rows, column=cols).value = countr
        rows += 1

    rows = 2
    cols = 2
    for kist in amount:
        sheet.cell(row= rows,column = cols).value = kist
        rows +=1

    wb.save('korona.xlsx')


    print('made!')

    # for i in range(len(amount)):
    #     print(all_country_arr[i])
    #     print(amount[i])

    # print(country.text)

    # print(country)

while True:
    x=datetime.today()
    y=x.replace(day=x.day, hour=22, minute=33, second=20, microsecond=0)
    delta_t=str(y-x)
    delta_t = delta_t[0:7]

    print(delta_t)

    if delta_t == '0:00:00':
        korona()
        break





